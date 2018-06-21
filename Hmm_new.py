
import pandas as pd
import numpy as np
from sklearn import hmm
import openpyxl
import warnings
warnings.filterwarnings("ignore")

JIM  = 1 # # of rows for 1st MAX_USER_ID users
MAX_USER_ID = 2293 ; # or number of user to run the test on


print "Opening the work book ... \n"
wb= openpyxl.load_workbook('10-90tky.xlsx') # load the .xlsx excel file

#print type(wb)
#print wb.get_sheet_names() # Name of all Sheets in the file
print "Getting the required sheet ... \n"
sheet=wb.get_sheet_by_name('train_tky') # access particular sheet in the fil
#print sheet


# this while loop basically finds the last row number of the valid users for testing, ie 
# if MAX_USER_ID = 3 , JIM will be equal to 376
print "Finding the required row number to cover all " , MAX_USER_ID , " users  ...\n"
while(True):
	try :
		y_1= int((sheet.cell(row=JIM,column=1).value))	
	except :
		JIM -=1
		break 
	if y_1 == MAX_USER_ID + 1  :
		JIM -=1 
		break 
	JIM += 1 


print "Total number of rows to be processed are ...",JIM,"\n"	 

''' Code for calculating probablity of each ven_cat_name'''


print "Calculating all the required probabilities ... \n" 


list_diff_ven = [] # list of all the different places 
list_day = [] 	   # list of all the days 	
list_user = []     # list of all the users

# x is row number 
for x in range (1,JIM):
    #place 
    y= str((sheet.cell(row=x,column=2).value)) # take 2nd column
    list_diff_ven.append(y)
    
    list_diff_ven1=set(list_diff_ven) # for removing duplicate value of coloumn 2
    list_diff_ven=list(list_diff_ven1)
    
    # day
    y_6= str((sheet.cell(row=x,column=3).value))
    #y_6=y_6[0:4]
    list_day.append(y_6)
    
    # id
    y_1= str((sheet.cell(row=x,column=1).value))
    # print y_1
    list_user.append(y_1)


print "Done 1 / 11\n"

set_day = set(list_day)
list_day = list(set_day)
set_users = set(list_user)
list_user = list(set_users)

list_users = [int(u) for u in list_user] 
list_users.sort()
list_user = [str(u) for u in list_users]


user_day = [] # user-day pair
for user in list_user:  # Creating list of pair of user and day
    for day in list_day:
        list_two = []
        list_two.append(user)
        list_two.append(day)
        user_day.append(list_two)       
print "Done 2 / 11\n"


user_day_dict = {} 
# Dictionary for each member    
dict1={}
dict12={}
lenth = len(user_day)

# dict1 has 0 for each place 
for place in list_diff_ven:
    dict1[place] = 0
   

print "Done 3 / 11\n"
    
#dict12 has a list with 3*7 , users*days 0's in it for every place 
for place in list_diff_ven:
    list1 = []
    for i in range(lenth):
        list1.append(0)
    dict12[place] = list1  # Dictionary for each restaurant


print "Done 4 / 11\n"

for place in list_diff_ven: # for each place 
    for x in range (1,JIM):	# and each entry in the sheet
    	
    	#get place , user_id and day
        y1= str((sheet.cell(row=x,column=1).value))#id
        y2= str((sheet.cell(row=x,column=2).value))#place
        y6= str((sheet.cell(row=x,column=3).value))#day
        
        y6=y6[0:5]
        
        if y2 == place:
            # get total times a particular place is visited.
            dict1[place] = dict1[place] + 1
            list_temp = []
            list_temp.append(y1) # user_id
            list_temp.append(y6) # day
            index1 = user_day.index(list_temp) # checking for user day existing in previously created user day list
            dict12[place][index1] = dict12[place][index1] + 1

print "Done 5 / 11\n"

# Calculating initial probablity
# count of all the visits 
Total_Sum = 0.0                  

for place in list_diff_ven:
    Total_Sum = Total_Sum + dict1[place]  


print "Done 6 / 11\n"
                      
#Now dict1[place] = global probability of sm1 going to that place 
for place in list_diff_ven:
    dict1[place] = float(dict1[place]) / Total_Sum  # Get initial probablity of each hidden state


print "Done 7 / 11\n"

# Calculating emission probality
for place in list_diff_ven:
    sum1=0.0
	
    #sum1 = sum(dict12[place])
    
    for j in range(lenth):
        sum1 = sum1 + dict12[place][j] # each place with its user_day pair
    
    for j in range(lenth):
        dict12[place][j] = float(dict12[place][j]) / sum1


print "Done 8 / 11\n"

# Get ready initial and emission probality for hmm code
start_prob_member=[] # overall probability of each place 
emission_probab = [] # set of probabilities for each user-day pair for each place 

for place in list_diff_ven:
    start_prob_member.append(dict1[place]) # Finding the initial probabality of each place(Hidden state)
    emission_probab.append(dict12[place]) # Finding the emission probablity of each place(Hidden state)


print "Done 9 / 11\n"

total_len = len(list_diff_ven)
tran_p = 1.0/ total_len

list_i = []

for i in list_diff_ven:
    list_i.append(tran_p)


print "Done 10 / 11\n"

trans_probality = []    
for j in list_diff_ven:
    trans_probality.append(list_i)
            

print "Done 11 / 11 \n"
   
''' Code for hidden markov model'''        

print "Training the model with the probabilities ... \n"

states = list_diff_ven # total number of hidden state
n_states = len(states)

observations = list_day
n_observations = len(observations)

start_probability = np.array(start_prob_member) # has nothing to do with users 
#print "start=", start_probability	

transition_probability = np.array(trans_probality)	# has nothing to do with users 
#print "transition", transition_probability

emission_probability = np.array(emission_probab)	# here there are users involved
#print "emmission", emission_probability


model = hmm.MultinomialHMM(n_components=n_states)	

model._set_startprob(start_probability)
model._set_transmat(transition_probability)
model._set_emissionprob(emission_probability)		# here there are users involved

# predict a sequence of hidden states based on visible states

LIST_OF_DICT_FOR_PANDAS_DATAFRAME = []

for y in range(0,MAX_USER_ID):
	user_id = y+1
	#print "\n\nUser " , user_id ;
	for x in range (1,7):
		
		DICT_FOR_PANDAS_DATAFRAME = {}
		DICT_FOR_PANDAS_DATAFRAME["user_id"] = 	user_id ;
		
		# i = user_day array index 	
		i = 7*y + x ;
		bob_says = [i]
		
		logprob, alice_hears = model.decode(bob_says, algorithm="viterbi")
				
		# x = Day array index
		bob_says = [x] 
	
		DAY = ", ".join(map(lambda X: observations[X], bob_says)) 	
		DICT_FOR_PANDAS_DATAFRAME["day"] = DAY ;
		
		PLACES = ", ".join(map(lambda X: states[X], alice_hears))
		DICT_FOR_PANDAS_DATAFRAME["place"] = PLACES ;	
	
		LIST_OF_DICT_FOR_PANDAS_DATAFRAME.append(DICT_FOR_PANDAS_DATAFRAME)
				
		#print "Bob says:", ", ".join(map(lambda X: observations[X], bob_says))
		#print "Alice hears:", ", ".join(map(lambda X: states[X], alice_hears))


excel_name ="out19_tky_"+str(MAX_USER_ID)+"_train_users.xlsx"

print "\n\nWritting the output to ",excel_name," ... " 
df = pd.DataFrame(LIST_OF_DICT_FOR_PANDAS_DATAFRAME)
df = df[['user_id','day','place']]
df.to_excel(excel_name,index=False)
 
